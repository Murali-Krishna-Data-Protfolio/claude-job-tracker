"""
sync_cloud.py — Cloud replication + mobile push notifications.

Telegraph  → permanent public URL, auto-updated daily, works on any device
ntfy.sh    → instant push notification when new jobs are added (free, no account)

First run:  python sync_cloud.py --setup
Daily run:  called automatically by job_tracker.py
"""

import json
import os
import sys
from datetime import date, datetime
from pathlib import Path

import requests

ENV_PATH = Path(__file__).parent / ".env"

# ── Config keys stored in .env ─────────────────────────────────────────────────
TELEGRAPH_TOKEN_KEY  = "TELEGRAPH_ACCESS_TOKEN"
TELEGRAPH_PATH_KEY   = "TELEGRAPH_PAGE_PATH"
NTFY_TOPIC_KEY       = "NTFY_TOPIC"

TELEGRAPH_API = "https://api.telegra.ph"
NTFY_BASE     = "https://ntfy.sh"


# ── .env helpers ──────────────────────────────────────────────────────────────

def _read_env() -> dict[str, str]:
    if not ENV_PATH.exists():
        return {}
    out = {}
    for line in ENV_PATH.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if line and not line.startswith("#") and "=" in line:
            k, _, v = line.partition("=")
            out[k.strip()] = v.strip()
    return out


def _write_env_key(key: str, value: str):
    env = _read_env()
    env[key] = value
    lines = ENV_PATH.read_text(encoding="utf-8").splitlines()
    updated = False
    for i, line in enumerate(lines):
        if line.strip().startswith(f"{key}=") or line.strip() == f"{key}=":
            lines[i] = f"{key}={value}"
            updated = True
            break
    if not updated:
        lines.append(f"{key}={value}")
    ENV_PATH.write_text("\n".join(lines) + "\n", encoding="utf-8")


# ── Telegraph helpers ──────────────────────────────────────────────────────────

def _telegraph_post(endpoint: str, **kwargs) -> dict:
    r = requests.post(f"{TELEGRAPH_API}/{endpoint}", json=kwargs, timeout=15)
    r.raise_for_status()
    data = r.json()
    if not data.get("ok"):
        raise RuntimeError(f"Telegraph API error: {data.get('error')}")
    return data["result"]


def _create_telegraph_account() -> str:
    result = _telegraph_post(
        "createAccount",
        short_name="MuraliJobs",
        author_name="Murali Krishna — Job Tracker",
        author_url="https://www.linkedin.com/search/results/jobs/?keywords=data+analyst+france",
    )
    token = result["access_token"]
    _write_env_key(TELEGRAPH_TOKEN_KEY, token)
    print(f"  Telegraph account created. Token saved to .env")
    return token


def _build_content(jobs: list[dict]) -> list[dict]:
    """Build Telegraph node tree from job list."""
    today = date.today().strftime("%B %d, %Y")
    nodes: list[dict] = []

    # Header info line
    nodes.append({
        "tag": "p",
        "children": [
            {"tag": "em", "children": [
                f"Auto-updated: {today}  •  {len(jobs)} English-speaking jobs in France  •  "
            ]},
            {"tag": "a",
             "attrs": {"href": "https://www.linkedin.com/jobs/search-results/?keywords=Jobs%20english%20in%20france"},
             "children": ["LinkedIn Search"]},
        ]
    })
    nodes.append({"tag": "hr"})

    # Group by category
    categories: dict[str, list[dict]] = {}
    for j in jobs:
        title = j.get("Title") or j.get("title") or ""
        cat = "Other"
        for kw in ["Data Analyst", "Data Engineer", "Business Analyst",
                   "BI Developer", "Analytics Engineer"]:
            if kw.lower() in title.lower():
                cat = kw
                break
        categories.setdefault(cat, []).append(j)

    STATUS_ICON = {
        "Applied": "📤", "Interview": "📅", "Offer": "🎉",
        "Rejected": "❌", "Saved": "🔖",
    }

    cat_order = ["Data Analyst", "Data Engineer", "Business Analyst",
                 "BI Developer", "Analytics Engineer", "Other"]

    for cat in cat_order:
        cat_jobs = categories.get(cat)
        if not cat_jobs:
            continue

        nodes.append({
            "tag": "h3",
            "children": [f"{cat}  ({len(cat_jobs)})"]
        })

        for j in cat_jobs:
            title   = j.get("Title")   or j.get("title",   "")
            company = j.get("Company") or j.get("company", "")
            loc     = j.get("Location")or j.get("location","")
            salary  = j.get("Salary")  or j.get("salary")  or ""
            url     = j.get("Apply_URL")or j.get("url",    "")
            status  = j.get("Status")  or j.get("status",  "Saved")
            icon    = STATUS_ICON.get(status, "🔖")

            title_node: dict = (
                {"tag": "a", "attrs": {"href": url}, "children": [f"{icon} {title}"]}
                if url and url.startswith("http")
                else {"tag": "strong", "children": [f"{icon} {title}"]}
            )

            detail_parts = [company]
            if loc:
                detail_parts.append(f"📍 {loc}")
            if salary and salary.lower() not in ("not specified", "none", ""):
                detail_parts.append(f"💰 {salary}")
            detail_parts.append(f"[{status}]")

            nodes.append({
                "tag": "p",
                "children": [
                    title_node,
                    {"tag": "br"},
                    "  ".join(detail_parts),
                ]
            })

        nodes.append({"tag": "hr"})

    if not jobs:
        nodes.append({"tag": "p", "children": ["No jobs tracked yet. Run job_tracker.py to populate."]})

    return nodes


def _create_or_update_page(token: str, page_path: str | None,
                            jobs: list[dict]) -> tuple[str, str]:
    """Creates a new page or edits the existing one. Returns (url, path)."""
    today = date.today().strftime("%B %d, %Y")
    title = f"Job DB — {today} | {len(jobs)} English roles in France"
    content = _build_content(jobs)

    if page_path:
        result = _telegraph_post(
            f"editPage/{page_path}",
            access_token=token,
            title=title,
            content=content,
            author_name="Murali Krishna",
        )
    else:
        result = _telegraph_post(
            "createPage",
            access_token=token,
            title=title,
            content=content,
            author_name="Murali Krishna",
            return_content=False,
        )
        _write_env_key(TELEGRAPH_PATH_KEY, result["path"])

    return result["url"], result["path"]


# ── ntfy.sh helpers ────────────────────────────────────────────────────────────

def _send_ntfy(topic: str, title: str, body: str,
               click_url: str | None = None, priority: str = "default"):
    headers: dict[str, str] = {
        "Title":    title,
        "Priority": priority,
        "Tags":     "briefcase",
    }
    if click_url:
        headers["Click"] = click_url
    try:
        r = requests.post(
            f"{NTFY_BASE}/{topic}",
            data=body.encode("utf-8"),
            headers=headers,
            timeout=10,
        )
        return r.status_code == 200
    except Exception as e:
        print(f"  [warn] ntfy failed: {e}")
        return False


# ── Public API ─────────────────────────────────────────────────────────────────

def load_jobs_from_excel() -> list[dict]:
    """Read all rows from the Jobs sheet."""
    from config import OUTPUT_PATH
    from openpyxl import load_workbook
    if not Path(OUTPUT_PATH).exists():
        return []
    wb = load_workbook(OUTPUT_PATH, read_only=True, data_only=True)
    if "Jobs" not in wb.sheetnames:
        return []
    ws = wb["Jobs"]
    cols = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    jobs = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            jobs.append(dict(zip(cols, row)))
    wb.close()
    return jobs


def setup_cloud() -> dict[str, str]:
    """
    First-time setup: creates Telegraph account + page, picks ntfy topic.
    Saves everything to .env. Call once with: python sync_cloud.py --setup
    """
    print("\n=== Cloud Setup ===")
    env = _read_env()

    # Telegraph account
    token = env.get(TELEGRAPH_TOKEN_KEY, "")
    if not token:
        print("\n[1/3] Creating Telegraph account...")
        token = _create_telegraph_account()
    else:
        print(f"[1/3] Telegraph token already in .env — reusing")

    # Telegraph page (create with empty content first)
    page_path = env.get(TELEGRAPH_PATH_KEY, "")
    if not page_path:
        print("\n[2/3] Creating Telegraph job page...")
        jobs = load_jobs_from_excel()
        url, page_path = _create_or_update_page(token, None, jobs)
        print(f"  Page created: {url}")
    else:
        url = f"https://telegra.ph/{page_path}"
        print(f"[2/3] Telegraph page already exists: {url}")

    # ntfy topic
    import hashlib, time
    topic = env.get(NTFY_TOPIC_KEY, "")
    if not topic:
        print("\n[3/3] Generating ntfy topic...")
        # deterministic but unique to this user
        topic = "murali_jobs_" + hashlib.md5(
            "gkmurali37@gmail.com".encode()
        ).hexdigest()[:8]
        _write_env_key(NTFY_TOPIC_KEY, topic)
        print(f"  Topic: {topic}")
    else:
        print(f"[3/3] ntfy topic already set: {topic}")

    # Send test notification
    print("\nSending test push notification...")
    ok = _send_ntfy(
        topic,
        title="Job Tracker - Setup Complete",
        body=f"Your job DB is live. {len(load_jobs_from_excel())} jobs tracked.",
        click_url=url,
    )
    print(f"  Notification sent: {ok}")

    print("\n=== Setup Complete ===")
    print(f"  Job DB URL : {url}")
    print(f"  ntfy topic : {NTFY_BASE}/{topic}")
    print(f"\n  On your phone:")
    print(f"    1. Install 'ntfy' app (iOS / Android)")
    print(f"    2. Subscribe to: {NTFY_BASE}/{topic}")
    print(f"    3. Bookmark: {url}")
    print()

    return {"telegraph_url": url, "ntfy_url": f"{NTFY_BASE}/{topic}"}


def sync_cloud(added: int, total: int, verbose: bool = True) -> str | None:
    """
    Sync cloud DB after a run. Called by job_tracker.py automatically.
    Returns the Telegraph URL.
    """
    env = _read_env()
    token     = env.get(TELEGRAPH_TOKEN_KEY, "")
    page_path = env.get(TELEGRAPH_PATH_KEY,  "")
    topic     = env.get(NTFY_TOPIC_KEY,      "")

    if not token:
        if verbose:
            print("  [cloud] Not set up yet. Run: python sync_cloud.py --setup")
        return None

    # Update Telegraph page
    jobs = load_jobs_from_excel()
    try:
        url, _ = _create_or_update_page(token, page_path or None, jobs)
        if verbose:
            print(f"  Telegraph page updated: {url}")
    except Exception as e:
        if verbose:
            print(f"  [warn] Telegraph update failed: {e}")
        url = f"https://telegra.ph/{page_path}" if page_path else None

    # Push notification via ntfy.sh
    if topic and added > 0:
        today = date.today().strftime("%b %d")
        body  = f"+{added} new English-speaking jobs added  •  Total: {total}"
        ok = _send_ntfy(
            topic,
            title=f"job_db loaded - {today}  |  +{added} new jobs",
            body=body,
            click_url=url,
            priority="high" if added >= 5 else "default",
        )
        if verbose:
            print(f"  Push notification sent: {ok}")
    elif topic and added == 0 and verbose:
        print("  No new jobs — push notification skipped")

    return url


if __name__ == "__main__":
    if "--setup" in sys.argv:
        setup_cloud()
    else:
        # Quick re-sync (useful to force-update the page)
        env = _read_env()
        if not env.get(TELEGRAPH_TOKEN_KEY):
            print("Not set up. Run: python sync_cloud.py --setup")
            sys.exit(1)
        jobs = load_jobs_from_excel()
        url, _ = _create_or_update_page(
            env[TELEGRAPH_TOKEN_KEY],
            env.get(TELEGRAPH_PATH_KEY),
            jobs,
        )
        print(f"Page updated: {url}")
