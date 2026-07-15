"""
sync_bookmarks.py — syncs the Chrome 'Jobs' bookmarks folder
from the Excel job database. Run after job_tracker.py or standalone.

Called automatically by job_tracker.py after each daily update.
"""

import hashlib
import json
import shutil
import time
from pathlib import Path

from openpyxl import load_workbook

from config import OUTPUT_PATH

BOOKMARKS_PATH = Path(
    r"C:\Users\gkmur\AppData\Local\Google\Chrome\User Data\Default\Bookmarks"
)
LINKEDIN_SEARCH = (
    "https://www.linkedin.com/jobs/search-results/"
    "?keywords=Jobs%20english%20in%20france&origin=SEMANTIC_SEARCH_LANDING_PAGE"
)

# Windows FILETIME offset (microseconds between 1601-01-01 and 1970-01-01)
_FT_OFFSET = 11_644_473_600_000_000


def _now_ft() -> str:
    return str(int(time.time() * 1_000_000) + _FT_OFFSET)


def _next_id(counter: list) -> str:
    v = counter[0]
    counter[0] += 1
    return str(v)


def _max_id(node: dict) -> int:
    m = int(node.get("id", 0))
    for child in node.get("children", []):
        m = max(m, _max_id(child))
    return m


def _compute_checksum(data: dict) -> str:
    md5 = hashlib.md5()
    def visit(node):
        md5.update(str(node["id"]).encode("utf-8"))
        if node["type"] == "url":
            md5.update(node["url"].encode("utf-8"))
        for child in node.get("children", []):
            visit(child)
    for key in ("bookmark_bar", "other", "synced"):
        visit(data["roots"][key])
    return md5.hexdigest()


def _url_node(title: str, url: str, id_counter: list) -> dict:
    return {
        "date_added": _now_ft(), "guid": "",
        "id": _next_id(id_counter),
        "name": title[:80],   # Chrome truncates long names
        "type": "url", "url": url,
    }


def _folder(name: str, children: list, id_counter: list) -> dict:
    return {
        "children": children,
        "date_added": _now_ft(), "date_modified": _now_ft(),
        "guid": "", "id": _next_id(id_counter),
        "name": name, "type": "folder",
    }


def read_jobs_from_excel() -> list[dict]:
    if not Path(OUTPUT_PATH).exists():
        return []
    wb = load_workbook(OUTPUT_PATH, read_only=True, data_only=True)
    if "Jobs" not in wb.sheetnames:
        return []
    ws = wb["Jobs"]
    jobs = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        url = row[7] or ""
        if not url or not url.startswith("http"):
            continue
        jobs.append({
            "title":   str(row[2] or "").strip(),
            "company": str(row[3] or "").strip(),
            "location":str(row[4] or "").strip(),
            "status":  str(row[6] or "Saved").strip(),
            "url":     url,
            "query":   str(row[10] or "Other").strip(),
        })
    wb.close()
    return jobs


def sync_bookmarks(verbose: bool = True) -> int:
    """
    Rebuild the Chrome 'Jobs' bookmarks folder from the Excel DB.
    Returns number of bookmarks written.
    """
    if not BOOKMARKS_PATH.exists():
        if verbose:
            print("  [bookmarks] Chrome Bookmarks file not found — skipped")
        return 0

    data = json.loads(BOOKMARKS_PATH.read_text(encoding="utf-8"))
    roots = data["roots"]

    # Find current max ID across all nodes
    cur_max = max(
        _max_id(roots["bookmark_bar"]),
        _max_id(roots["other"]),
        _max_id(roots["synced"]),
        int(roots["bookmark_bar"].get("id", "1")),
    )
    counter = [cur_max + 1]

    # Read jobs
    all_jobs = read_jobs_from_excel()

    # Group by category
    STATUS_EMOJI = {
        "Applied":   "Applied",
        "Interview": "Interview",
        "Offer":     "Offer",
        "Rejected":  "Rejected",
        "Saved":     "",
    }

    categories: dict[str, list[dict]] = {}
    for j in all_jobs:
        cat = "Other"
        for kw in ["Data Analyst", "Data Engineer", "Business Analyst",
                   "BI Developer", "Analytics Engineer"]:
            if kw.lower() in j["title"].lower():
                cat = kw
                break
        categories.setdefault(cat, []).append(j)

    # Build Jobs folder contents
    children = []

    # 1. LinkedIn search shortcut at the top
    children.append(_url_node(
        "LinkedIn — English jobs in France (search)", LINKEDIN_SEARCH, counter
    ))

    # 2. Category sub-folders
    for cat_name in ["Data Analyst", "Data Engineer", "Business Analyst",
                     "BI Developer", "Analytics Engineer", "Other"]:
        jobs_in_cat = categories.get(cat_name, [])
        if not jobs_in_cat:
            continue
        nodes = []
        for j in jobs_in_cat:
            status_tag = STATUS_EMOJI.get(j["status"], "")
            prefix = f"[{status_tag}] " if status_tag else ""
            label = f"{prefix}{j['title']} @ {j['company']} — {j['location']}"
            nodes.append(_url_node(label, j["url"], counter))
        children.append(_folder(
            f"{cat_name} ({len(nodes)})", nodes, counter
        ))

    jobs_folder = _folder(
        f"Jobs ({len(all_jobs)} total)", children, counter
    )

    # Replace existing Jobs folder or prepend
    bar = roots["bookmark_bar"]
    bar.setdefault("children", [])
    bar["children"] = [
        c for c in bar["children"] if c.get("name", "").split(" (")[0] != "Jobs"
    ]
    bar["children"].insert(0, jobs_folder)
    bar["date_modified"] = _now_ft()

    # Recompute checksum
    data["checksum"] = _compute_checksum(data)

    # Backup + write
    shutil.copy2(BOOKMARKS_PATH, BOOKMARKS_PATH.with_suffix(".bak_claude"))
    BOOKMARKS_PATH.write_text(
        json.dumps(data, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )

    total = len(all_jobs)
    if verbose:
        print(f"  Chrome 'Jobs' folder synced: {total} bookmarks across {len(categories)} categories")
        print("  Restart Chrome or press Ctrl+Shift+B to see changes")
    return total


if __name__ == "__main__":
    sync_bookmarks()
