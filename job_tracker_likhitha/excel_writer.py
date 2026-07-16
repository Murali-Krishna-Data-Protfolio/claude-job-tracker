"""
Creates and updates the job_applications.xlsx with:
  - Sheet "Jobs"      : structured data table (Power BI-ready)
  - Sheet "Dashboard" : KPI cells + charts built with openpyxl
"""

import os
from datetime import date, datetime
from typing import Any

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    GradientFill,
    PatternFill,
    Side,
    numbers,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

from config import OUTPUT_PATH, STATUS_CHOICES

# ── Colour palette (Murali's Power BI / dark-blue theme) ──────────────────────
C_HEADER_BG   = "1F3864"   # dark navy
C_HEADER_FG   = "FFFFFF"
C_ALT_ROW     = "EAF0FB"   # light blue-grey
C_ACCENT      = "2E75B6"   # mid blue
C_GREEN       = "70AD47"
C_ORANGE      = "ED7D31"
C_RED         = "FF0000"
C_YELLOW      = "FFD966"

COLUMNS = [
    "Date_Found",
    "Job_ID",
    "Title",
    "Company",
    "Location",
    "Salary",
    "Status",
    "Apply_URL",
    "Description_Preview",
    "English_Confidence",
    "Search_Query",
    "English_Reason",
]

COL_WIDTHS = {
    "Date_Found": 14,
    "Job_ID": 18,
    "Title": 38,
    "Company": 28,
    "Location": 22,
    "Salary": 18,
    "Status": 14,
    "Apply_URL": 40,
    "Description_Preview": 55,
    "English_Confidence": 20,
    "Search_Query": 20,
    "English_Reason": 40,
}


def _header_style(cell, bg: str = C_HEADER_BG, fg: str = C_HEADER_FG):
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(bold=True, color=fg, size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _thin_border():
    side = Side(style="thin", color="BFBFBF")
    return Border(left=side, right=side, top=side, bottom=side)


def _load_or_create() -> tuple[Workbook, bool]:
    """Return (workbook, is_new)."""
    if os.path.exists(OUTPUT_PATH):
        return load_workbook(OUTPUT_PATH), False
    wb = Workbook()
    # remove default sheet
    wb.remove(wb.active)
    return wb, True


def _ensure_jobs_sheet(wb: Workbook) -> openpyxl.worksheet.worksheet.Worksheet:
    if "Jobs" not in wb.sheetnames:
        ws = wb.create_sheet("Jobs", 0)
        _write_jobs_header(ws)
    return wb["Jobs"]


def _write_jobs_header(ws):
    for i, col in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=i, value=col)
        _header_style(cell)
        ws.column_dimensions[get_column_letter(i)].width = COL_WIDTHS.get(col, 20)

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"


def load_existing_job_ids(wb: Workbook) -> set[str]:
    if "Jobs" not in wb.sheetnames:
        return set()
    ws = wb["Jobs"]
    ids = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1]:  # Job_ID is column index 1
            ids.add(str(row[1]))
    return ids


def append_jobs(wb: Workbook, jobs: list[dict]) -> int:
    """Append new jobs to the Jobs sheet. Returns count of rows added."""
    ws = _ensure_jobs_sheet(wb)
    existing_ids = load_existing_job_ids(wb)

    added = 0
    for job in jobs:
        jid = str(job.get("job_id", ""))
        if jid in existing_ids:
            continue

        row_num = ws.max_row + 1
        is_alt = row_num % 2 == 0
        fill = PatternFill("solid", fgColor=C_ALT_ROW) if is_alt else None

        values = [
            date.today().isoformat(),
            jid,
            job.get("title", ""),
            job.get("company", ""),
            job.get("location", ""),
            job.get("salary") or "Not specified",
            "Saved",
            job.get("url", ""),
            (job.get("description", "")[:300] + "...") if len(job.get("description", "")) > 300 else job.get("description", ""),
            job.get("english_confidence", ""),
            job.get("search_query", ""),
            job.get("english_reason", ""),
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.border = _thin_border()
            cell.alignment = Alignment(vertical="center", wrap_text=(col_idx in (3, 9, 12)))
            if fill:
                cell.fill = fill

            # Hyperlink for Apply_URL column
            if col_idx == 8 and value and value.startswith("http"):
                cell.hyperlink = value
                cell.font = Font(color="0563C1", underline="single")

        existing_ids.add(jid)
        added += 1

    # Add/refresh Excel Table
    if ws.max_row >= 2:
        last_col = get_column_letter(len(COLUMNS))
        ref = f"A1:{last_col}{ws.max_row}"
        # Remove existing tables first
        for tbl in list(ws.tables.values()):
            del ws.tables[tbl.name]
        table = Table(displayName="JobsTable", ref=ref)
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        ws.add_table(table)

    # Status dropdown validation
    if ws.max_row > 1:
        dv = DataValidation(
            type="list",
            formula1=f'"{",".join(STATUS_CHOICES)}"',
            showDropDown=False,
        )
        ws.add_data_validation(dv)
        dv.add(f"G2:G{ws.max_row}")

    return added


def _build_dashboard(wb: Workbook):
    """Rebuild the Dashboard sheet with KPI summary and charts."""
    if "Dashboard" in wb.sheetnames:
        del wb["Dashboard"]

    ws_jobs = wb["Jobs"] if "Jobs" in wb.sheetnames else None
    ws = wb.create_sheet("Dashboard")

    # ── Title ──────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:J1")
    title_cell = ws["A1"]
    title_cell.value = "Job Application Tracker — BI Dashboard"
    title_cell.font = Font(bold=True, size=16, color=C_HEADER_FG)
    title_cell.fill = PatternFill("solid", fgColor=C_HEADER_BG)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:J2")
    ws["A2"].value = f"Last refreshed: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(italic=True, size=10, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")

    if ws_jobs is None or ws_jobs.max_row < 2:
        ws["A4"].value = "No job data yet. Run job_tracker.py to populate."
        return

    # ── Read data from Jobs sheet ──────────────────────────────────────────────
    jobs_data = []
    for row in ws_jobs.iter_rows(min_row=2, values_only=True):
        if row[0]:
            jobs_data.append(dict(zip(COLUMNS, row)))

    if not jobs_data:
        return

    total = len(jobs_data)
    by_status: dict[str, int] = {}
    by_title: dict[str, int] = {}
    by_company: dict[str, int] = {}
    by_date: dict[str, int] = {}

    for j in jobs_data:
        status = j.get("Status") or "Saved"
        by_status[status] = by_status.get(status, 0) + 1

        title = j.get("Title") or "Unknown"
        # Normalise to broad category
        cat = "Other"
        for kw in ["Data Analyst", "Data Engineer", "Business Analyst", "BI Developer", "Analytics Engineer"]:
            if kw.lower() in title.lower():
                cat = kw
                break
        by_title[cat] = by_title.get(cat, 0) + 1

        company = j.get("Company") or "Unknown"
        by_company[company] = by_company.get(company, 0) + 1

        d = str(j.get("Date_Found") or "")[:10]
        if d:
            by_date[d] = by_date.get(d, 0) + 1

    applied = by_status.get("Applied", 0)
    interviews = by_status.get("Interview", 0)
    offers = by_status.get("Offer", 0)

    # ── KPI Cards (row 4–6) ────────────────────────────────────────────────────
    kpis = [
        ("Total Jobs Found", total, C_ACCENT),
        ("Applied", applied, C_GREEN),
        ("Interviews", interviews, C_ORANGE),
        ("Offers", offers, C_YELLOW),
    ]
    kpi_cols = [1, 3, 5, 7]
    for (label, value, color), col in zip(kpis, kpi_cols):
        ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col + 1)
        ws.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col + 1)
        ws.merge_cells(start_row=6, start_column=col, end_row=6, end_column=col + 1)

        lbl_cell = ws.cell(row=4, column=col, value=label)
        lbl_cell.font = Font(bold=True, size=10, color="FFFFFF")
        lbl_cell.fill = PatternFill("solid", fgColor=color)
        lbl_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[4].height = 20

        val_cell = ws.cell(row=5, column=col, value=value)
        val_cell.font = Font(bold=True, size=22, color=color)
        val_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[5].height = 34

        ws.cell(row=6, column=col).fill = PatternFill("solid", fgColor="F2F2F2")

    # ── Data tables for charts ─────────────────────────────────────────────────
    # Table 1: Jobs by Role (col A, starting row 9)
    ws.cell(row=8, column=1, value="Jobs by Role").font = Font(bold=True, color=C_HEADER_FG)
    ws.cell(row=8, column=1).fill = PatternFill("solid", fgColor=C_ACCENT)
    ws.cell(row=8, column=2, value="Count").font = Font(bold=True, color=C_HEADER_FG)
    ws.cell(row=8, column=2).fill = PatternFill("solid", fgColor=C_ACCENT)
    role_start = 9
    for i, (role, cnt) in enumerate(sorted(by_title.items(), key=lambda x: -x[1])):
        ws.cell(row=role_start + i, column=1, value=role)
        ws.cell(row=role_start + i, column=2, value=cnt)
    role_end = role_start + len(by_title) - 1

    # Table 2: Application Status (col D)
    ws.cell(row=8, column=4, value="Status").font = Font(bold=True, color=C_HEADER_FG)
    ws.cell(row=8, column=4).fill = PatternFill("solid", fgColor=C_ACCENT)
    ws.cell(row=8, column=5, value="Count").font = Font(bold=True, color=C_HEADER_FG)
    ws.cell(row=8, column=5).fill = PatternFill("solid", fgColor=C_ACCENT)
    status_start = 9
    for i, (status, cnt) in enumerate(sorted(by_status.items(), key=lambda x: -x[1])):
        ws.cell(row=status_start + i, column=4, value=status)
        ws.cell(row=status_start + i, column=5, value=cnt)
    status_end = status_start + len(by_status) - 1

    # Table 3: Top Companies (col G)
    ws.cell(row=8, column=7, value="Company").font = Font(bold=True, color=C_HEADER_FG)
    ws.cell(row=8, column=7).fill = PatternFill("solid", fgColor=C_ACCENT)
    ws.cell(row=8, column=8, value="Jobs").font = Font(bold=True, color=C_HEADER_FG)
    ws.cell(row=8, column=8).fill = PatternFill("solid", fgColor=C_ACCENT)
    top_companies = sorted(by_company.items(), key=lambda x: -x[1])[:10]
    comp_start = 9
    for i, (comp, cnt) in enumerate(top_companies):
        ws.cell(row=comp_start + i, column=7, value=comp)
        ws.cell(row=comp_start + i, column=8, value=cnt)
    comp_end = comp_start + len(top_companies) - 1

    # ── Chart 1: Jobs by Role (BarChart) ──────────────────────────────────────
    chart1 = BarChart()
    chart1.type = "col"
    chart1.title = "Jobs by Role"
    chart1.y_axis.title = "Count"
    chart1.style = 10
    chart1.width = 16
    chart1.height = 12
    chart1.grouping = "clustered"

    data_ref = Reference(ws, min_col=2, min_row=8, max_row=role_end)
    cats_ref = Reference(ws, min_col=1, min_row=role_start, max_row=role_end)
    chart1.add_data(data_ref, titles_from_data=True)
    chart1.set_categories(cats_ref)
    chart1.series[0].graphicalProperties.solidFill = C_ACCENT
    ws.add_chart(chart1, "A20")

    # ── Chart 2: Status breakdown (PieChart) ──────────────────────────────────
    chart2 = PieChart()
    chart2.title = "Application Status"
    chart2.style = 10
    chart2.width = 16
    chart2.height = 12

    data_ref2 = Reference(ws, min_col=5, min_row=8, max_row=status_end)
    cats_ref2 = Reference(ws, min_col=4, min_row=status_start, max_row=status_end)
    chart2.add_data(data_ref2, titles_from_data=True)
    chart2.set_categories(cats_ref2)
    ws.add_chart(chart2, "E20")

    # ── Chart 3: Top Companies (BarChart horizontal) ───────────────────────────
    chart3 = BarChart()
    chart3.type = "bar"
    chart3.title = "Top Hiring Companies"
    chart3.y_axis.title = "Jobs"
    chart3.style = 10
    chart3.width = 16
    chart3.height = 12

    data_ref3 = Reference(ws, min_col=8, min_row=8, max_row=comp_end)
    cats_ref3 = Reference(ws, min_col=7, min_row=comp_start, max_row=comp_end)
    chart3.add_data(data_ref3, titles_from_data=True)
    chart3.set_categories(cats_ref3)
    chart3.series[0].graphicalProperties.solidFill = C_GREEN
    ws.add_chart(chart3, "I20")

    # Column widths
    for col, width in [(1, 22), (2, 10), (3, 2), (4, 16), (5, 10), (6, 2), (7, 28), (8, 10)]:
        ws.column_dimensions[get_column_letter(col)].width = width


def save_workbook(wb: Workbook):
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"\n  Saved: {OUTPUT_PATH}")
